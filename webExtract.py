from Tkinter import *
import sys
import requests 
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

root = Tk()

wb = Workbook()
ws1 = wb.active
ws1.title = "Indeed"


def searchWebSite():
	job = jobEntry.get()
	job = job.replace(" ", "+")

	city = cityEntry.get()
	city = city.replace( " ", "+")

	state = stateEntry.get()
	state = state.replace(" ", "+")

	url = "http://www.indeed.com/jobs?q=" + job + "&l=" + city + "%2C+" + state
	sourceCode = requests.get(url)
	plainText = sourceCode.text
	soup = BeautifulSoup(plainText)
	jobNames = soup.find_all("h2", {"class": "jobtitle"})
	jobCompanies = soup.find_all("span", {"itemprop": "name"})
	jobLocation = soup.find_all("span", {"itemprop": "addressLocality"})
	jobLink = soup.find_all("a", {"itemprop": "title"})
	listNames = []
	listCompanies = []
	listLocations = []
	listLinks = []
	listEverything = []
	rowNum = 0

		
	for item in jobNames:
		listNames.append(item.text)
		listEverything.append(item.text)
	
	
	for item in jobCompanies:
		listCompanies.append(item.text)
		listEverything.append(item.text)
	
	for item in jobLocation:
		listLocations.append(item.text)
		listEverything.append(item.text)
	
	for link in jobLink:
		href = link.get("href")
		listLinks.append("www.indeed.com" + href)
		listEverything.append("www.indeed.com" + href)

	totalListAmount = len(listEverything)
	totalNamesAmount = len(listNames)
	totalCompaniesAmount = len(listCompanies)
	totalLocationsAmount = len(listLocations)
	totalLinksAmount = len(listLinks)
	curRow = 1

	
	ws1.cell(row = curRow, column = 1).value = 'JOB DESCRIPTION'
	curRow += 1
	while(totalNamesAmount > 0):
			listEverything[rowNum] = listEverything[rowNum].strip()
			ws1.cell(row = curRow, column = 1).value = listEverything[rowNum] 
			totalNamesAmount += -1
			rowNum += 1
			curRow += 1

	

	curRow = 1
	ws1.cell(row = curRow, column = 7).value = 'COMPANY'
	curRow += 1
	while(totalCompaniesAmount > 0):
			ws1.cell(row = curRow, column = 7).value = listEverything[rowNum]
			totalCompaniesAmount += -1
			rowNum += 1
			curRow += 1

	curRow = 1
	ws1.cell(row = curRow, column = 11).value = 'LOCATION'
	curRow += 1
	while(totalLocationsAmount > 0):
			ws1.cell(row = curRow, column = 11).value = listEverything[rowNum]
			totalLocationsAmount += -1
			rowNum += 1
			curRow += 1


	curRow = 1
	ws1.cell(row = curRow, column  = 14).value = 'Links'
	curRow += 1
	while(totalLinksAmount > 0):
			ws1.cell(row = curRow, column  = 14).value = listEverything[rowNum]
			totalLinksAmount += -1
			rowNum += 1
			curRow += 1

	wb.save('jobList.xlsx')

##top frame
topFrame = Frame(root)
topFrame.pack(side=TOP)


##bottom frame
bottomFrame = Frame(root)
bottomFrame.pack(side=BOTTOM)


##label that tells user what to enter
label1 = Label(topFrame, text="Job Search Excel Creater")
label1.pack(side=TOP)


##label two
label2 = Label(topFrame, text="Job:")
label2.pack(side=LEFT)


##job entry text box
jobEntry = Entry(topFrame)
jobEntry.pack(side=LEFT)

##label three
label3 = Label(topFrame, text="State:")
label3.pack(side=LEFT)

##state entry text box
stateEntry = Entry(topFrame)
stateEntry.pack(side=LEFT)

##label four
label4 = Label(topFrame, text="City:")
label4.pack(side=LEFT)

##city entry text box
cityEntry = Entry(topFrame)
cityEntry.pack(side=LEFT)


##button that activates searchWebSite function
button1 = Button(bottomFrame, text="Create", command=searchWebSite)
button1.pack()

root.mainloop()
